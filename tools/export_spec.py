"""Generate the machine-readable spec from the document generators.

    python tools/export_spec.py

Writes spec/features.yaml and spec/evals.yaml.

WHY THIS EXISTS
---------------
"Did we build what the PRD says" is only answerable if the PRD is readable by a
program. A Word document is not. So the feature contracts are captured from the
same generator that renders the document, and the evals from the same generator
that renders the plan. Nothing is retyped, so nothing can drift.

The invariant this file protects: regenerating must produce no diff. If it does,
a generator changed and the spec was not refreshed -- `tools/trace.py` treats
that as a failure, not a warning.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.exit("pyyaml is required: pip install pyyaml")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
PRD_DIR = ROOT / "spec" / "prd"
PLAN_XLSX = ROOT / "docs" / "Nyaymalaw_Project_Plan.xlsx"
FEATURES_OUT = ROOT / "spec" / "features.yaml"
EVALS_OUT = ROOT / "spec" / "evals.yaml"

STATUSES = ("decided", "built", "tested", "verified live")


def features_from_prd() -> list[dict]:
    """Run the PRD generator in capture mode and read back the contracts."""
    proc = subprocess.run(
        ["node", "export_features.js"],
        cwd=PRD_DIR, capture_output=True, text=True, encoding="utf8",
    )
    if proc.returncode != 0:
        sys.exit(f"PRD generator failed:\n{proc.stderr[:2000]}")
    return json.loads(proc.stdout)


def _rows(ws, header_row: int) -> list[dict]:
    headers = [c.value for c in ws[header_row]]
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            continue
        out.append({h: v for h, v in zip(headers, row) if h})
    return out


def _header_row(ws) -> int:
    """The sheets carry a title and a note above the header; find the header."""
    for r in range(1, 6):
        if ws.cell(r, 1).value == "ID" or ws.cell(r, 1).value == "Feature":
            return r
    raise SystemExit(f"could not locate header row in sheet {ws.title!r}")


def plan_tables() -> tuple[list[dict], list[dict], list[dict]]:
    if not PLAN_XLSX.exists():
        sys.exit(f"missing {PLAN_XLSX} -- run spec/plan/build_plan.py first")
    wb = openpyxl.load_workbook(PLAN_XLSX, read_only=False, data_only=True)
    fmap = _rows(wb["Feature Map"], _header_row(wb["Feature Map"]))
    evals = _rows(wb["Evals"], _header_row(wb["Evals"]))
    tasks = _rows(wb["Tasks"], _header_row(wb["Tasks"]))
    return fmap, evals, tasks


def main() -> int:
    contracts = features_from_prd()
    fmap, evals, tasks = plan_tables()

    by_id = {f["Feature"]: f for f in fmap if f.get("Feature")}

    # A missing column must not read as "no evals". Resolve the header once and
    # fail loudly if it is absent -- an absent input silently producing an empty
    # list is defect shape S1, and it produced 43 eval-less features on the
    # first run of this exporter.
    eval_col = next((c for c in (fmap[0] if fmap else {}) if str(c).startswith("Evals")), None)
    if eval_col is None:
        sys.exit("Feature Map has no 'Evals...' column -- refusing to emit an "
                 "eval-less spec. Check spec/plan/build_plan.py.")
    tasks_by_feature: dict[str, list[str]] = {}
    for t in tasks:
        ref = str(t.get("PRD ref") or "")
        for fid in by_id:
            if fid in [r.strip() for r in ref.replace(";", ",").split(",")]:
                tasks_by_feature.setdefault(fid, []).append(t["ID"])

    features = []
    orphans = []
    for c in contracts:
        meta = by_id.get(c["id"])
        if meta is None:
            orphans.append(c["id"])
        eval_ids = []
        raw = meta.get(eval_col) if meta else None
        if raw and str(raw).strip() not in ("—", "-", ""):
            eval_ids = [e.strip() for e in str(raw).split(",") if e.strip()]
        features.append({
            "id": c["id"],
            "title": c["title"],
            "phase": (meta or {}).get("Phase"),
            "slice": (meta or {}).get("Slice"),
            "status": (meta or {}).get("Status", "decided"),
            "does": c["does"],
            "never": c["never"],
            "produces": c["produces"],
            "eval_prose": c["evals"],
            "eval_ids": eval_ids,
            "counterexample": c["counterexample"],
            "tasks": sorted(tasks_by_feature.get(c["id"], [])),
        })

    eval_rows = [{
        "id": e["ID"],
        "slice": e.get("Slice"),
        "class": e.get("Class"),
        "asserts": e.get("What it asserts"),
        "counterexample": e.get("The counterexample it MUST reject"),
        "cadence": e.get("Cadence"),
        "automated": e.get("Automated"),
        "prd_ref": e.get("PRD ref"),
    } for e in evals if e.get("ID")]

    FEATURES_OUT.write_text(
        "# GENERATED by tools/export_spec.py -- do not edit.\n"
        "# Source: spec/prd/*.js (the same generator that renders the Word document).\n"
        "# Regenerate after any change to the PRD generator, or trace.py will fail.\n\n"
        + yaml.safe_dump({"features": features}, sort_keys=False, allow_unicode=True, width=100),
        encoding="utf8")

    EVALS_OUT.write_text(
        "# GENERATED by tools/export_spec.py -- do not edit.\n"
        "# Source: spec/plan/build_plan.py via docs/Nyaymalaw_Project_Plan.xlsx.\n\n"
        + yaml.safe_dump({"evals": eval_rows}, sort_keys=False, allow_unicode=True, width=100),
        encoding="utf8")

    print(f"features : {len(features):>3}  -> {FEATURES_OUT.relative_to(ROOT)}")
    print(f"evals    : {len(eval_rows):>3}  -> {EVALS_OUT.relative_to(ROOT)}")
    by_status: dict[str, int] = {}
    for f in features:
        by_status[f["status"]] = by_status.get(f["status"], 0) + 1
    print("status   :", dict(sorted(by_status.items())))
    if orphans:
        print(f"\nWARNING: {len(orphans)} feature(s) in the PRD with no Feature Map row:")
        for o in orphans:
            print("   ", o)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
